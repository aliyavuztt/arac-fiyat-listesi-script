"""Excel çıktı servisi — marka bazlı ve birleşik dosya üretimi."""
import os
from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from Models.vehicle_price_record import VehiclePriceRecord


# Sayı formatı olan kolonlar (Excel header ismine göre).
# '#,##0' Excel'in yerel ayar duyarlı binlik ayırıcı formatıdır; Türkçe Excel'de
# '1.535.000' olarak görüntülenir ama hücre gerçek sayı kalır -> SUM çalışır.
NUMERIC_FORMAT = "#,##0"
NUMERIC_HEADERS = {"TavsiyeEdilenFiyat", "KampanyaliFiyat", "OpsiyonFiyati"}


class ExcelExportService:
    """docs/ klasörüne tek çalışma sayfalı .xlsx dosyaları yazar."""

    def __init__(self, docs_folder: str = "docs"):
        self.docs_folder = docs_folder

    def export(self, filename: str, records: List[VehiclePriceRecord]) -> str:
        """Kayıtları tek sayfaya yaz. Tam dosya yolunu döndür."""
        os.makedirs(self.docs_folder, exist_ok=True)
        filepath = os.path.join(self.docs_folder, filename)

        wb = Workbook()
        ws = wb.active
        ws.title = "Fiyat Listesi"

        headers = VehiclePriceRecord.excel_headers()
        ws.append(headers)
        self._style_header(ws, len(headers))

        # Kayıtları gruplu sırala: her versiyon+fiyat ana satırı hemen
        # ardından o gruba ait opsiyon satırları gelir.
        ordered = self._group_records(records)

        for rec in ordered:
            ws.append(rec.to_row())

        self._format_numeric_cells(ws, headers)
        self._auto_width(ws, len(headers))

        wb.save(filepath)
        return os.path.abspath(filepath)

    @staticmethod
    def _group_key(rec: VehiclePriceRecord) -> Tuple:
        """Bir kaydın grup anahtarını döndür.

        Aynı (marka, model, versiyon, tavsiye_edilen_fiyat, kampanyali_fiyat)
        değerine sahip ana satır ve opsiyon satırları aynı gruba düşer.
        Opel'de bir versiyonun birden çok fiyatı varsa her fiyat ayrı grup
        olur; her blok kendi opsiyonlarını taşır.
        """
        return (
            rec.marka,
            rec.model,
            rec.versiyon,
            rec.tavsiye_edilen_fiyat,
            rec.kampanyali_fiyat,
        )

    def _group_records(
        self, records: List[VehiclePriceRecord]
    ) -> List[VehiclePriceRecord]:
        """Kayıtları gruplu sırala: ana satır + o grubun opsiyonları.

        Strateji:
          1. Her grup için ilk görülme sırasını hatırla (kaynak sırası).
          2. Her grup içinde: ana satır (OpsiyonAdi boş) önce, ardından
             opsiyon satırları kaynak sırasına göre.
          3. Grupları ilk görülme sırasına göre diz.

        Sonuç: scraper'ın ürettiği marka/model/versiyon sırası korunur,
        sadece opsiyonlar ilgili ana satırın hemen altına taşınır.
        """
        # Grup anahtarı -> [ilk_görülme_indeksi, ana_satır, [opsiyon_satırları]]
        groups: Dict[Tuple, dict] = {}
        for idx, rec in enumerate(records):
            key = self._group_key(rec)
            if key not in groups:
                groups[key] = {
                    "first_idx": idx,
                    "main": None,
                    "options": [],
                }
            if rec.opsiyon_adi:
                groups[key]["options"].append(rec)
            else:
                # Ana satır — yoksa ilk görüleni sakla
                if groups[key]["main"] is None:
                    groups[key]["main"] = rec
                else:
                    # Aynı grupta birden fazla ana satır varsa (nadiren),
                    # ekstra ana satırı opsiyon listesine eklemeden ayrı sakla.
                    # Pratikte olmamalı; güvenlik için ilk ana satırı koru.
                    pass

        # Grupları ilk görülme sırasına göre sırala
        ordered_keys = sorted(groups.keys(), key=lambda k: groups[k]["first_idx"])

        result: List[VehiclePriceRecord] = []
        for key in ordered_keys:
            g = groups[key]
            # Ana satır önce (varsa); yoksa (sadece opsiyon varsa) atla ve
            # opsiyonları direkt yaz
            if g["main"] is not None:
                result.append(g["main"])
            result.extend(g["options"])

        return result

    def _format_numeric_cells(self, ws, headers: List[str]) -> None:
        """Fiyat kolonlarındaki sayı hücrelerine binlik ayırıcı format uygula."""
        numeric_cols = {
            i + 1 for i, h in enumerate(headers) if h in NUMERIC_HEADERS
        }
        if not numeric_cols:
            return
        for row in range(2, ws.max_row + 1):
            for col in numeric_cols:
                cell = ws.cell(row=row, column=col)
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = NUMERIC_FORMAT

    def _style_header(self, ws, col_count: int):
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="1F4E78", end_color="1F4E78", fill_type="solid"
        )
        for col in range(1, col_count + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _auto_width(self, ws, col_count: int):
        for col in range(1, col_count + 1):
            max_len = len(str(ws.cell(row=1, column=col).value or ""))
            for row in range(2, ws.max_row + 1):
                val = ws.cell(row=row, column=col).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            letter = ws.cell(row=1, column=col).column_letter
            ws.column_dimensions[letter].width = min(max_len + 2, 50)
